"""
Excel Export Service for HuiManager
Generates Excel reports for various data exports
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


def format_currency(amount):
    """Format amount as Vietnamese currency"""
    if amount is None:
        return "0 đ"
    return f"{amount:,.0f}".replace(",", ".") + " đ"


def format_date(dt):
    """Format datetime to Vietnamese format"""
    if dt is None:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
        except:
            return dt
    return dt.strftime("%d/%m/%Y %H:%M")


def create_styled_workbook():
    """Create a workbook with common styles"""
    wb = Workbook()
    return wb


def apply_header_style(cell):
    """Apply header styling to a cell"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_data_style(cell, is_currency=False, is_date=False):
    """Apply data cell styling"""
    cell.alignment = Alignment(horizontal="right" if is_currency else "left", vertical="center")
    cell.border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )


def auto_adjust_columns(ws):
    """Auto-adjust column widths"""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(length + 2, 50)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


def generate_members_excel(members_data, hui_group_name=None):
    """
    Generate Excel file for members list
    
    Args:
        members_data: List of member dictionaries
        hui_group_name: Optional group name for title
    
    Returns:
        BytesIO buffer containing Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Danh sách thành viên"
    
    # Title
    title = f"DANH SÁCH THÀNH VIÊN"
    if hui_group_name:
        title += f" - {hui_group_name}"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color="0369A1")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Export date
    ws['A2'] = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="64748B")
    
    # Headers
    headers = ["STT", "Họ tên", "Số điện thoại", "Số chân", "Điểm tín dụng", "Mức độ rủi ro", "Trạng thái", "Ghi chú"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Data
    risk_level_map = {
        'low': 'Thấp',
        'medium': 'Trung bình',
        'high': 'Cao',
        'critical': 'Rất cao'
    }
    
    for idx, member in enumerate(members_data, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=member.get('name', ''))
        ws.cell(row=row, column=3, value=member.get('phone', ''))
        ws.cell(row=row, column=4, value=member.get('slot_count', 1))
        ws.cell(row=row, column=5, value=member.get('credit_score', 100))
        ws.cell(row=row, column=6, value=risk_level_map.get(member.get('risk_level', 'low'), 'Thấp'))
        ws.cell(row=row, column=7, value='Hoạt động' if member.get('is_active', True) else 'Ngừng')
        ws.cell(row=row, column=8, value=member.get('notes', ''))
        
        for col in range(1, 9):
            apply_data_style(ws.cell(row=row, column=col))
    
    auto_adjust_columns(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_payments_excel(payments_data, hui_group_name=None, cycle_number=None):
    """
    Generate Excel file for payment history
    
    Args:
        payments_data: List of payment dictionaries
        hui_group_name: Optional group name for title
        cycle_number: Optional cycle number
    
    Returns:
        BytesIO buffer containing Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Lịch sử thanh toán"
    
    # Title
    title = "LỊCH SỬ THANH TOÁN"
    if hui_group_name:
        title += f" - {hui_group_name}"
    if cycle_number:
        title += f" - Kỳ {cycle_number}"
    ws.merge_cells('A1:I1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color="0369A1")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Export date
    ws['A2'] = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="64748B")
    
    # Headers
    headers = ["STT", "Thành viên", "Số tiền", "Mã tham chiếu", "Phương thức", "Trạng thái", "Hạn đóng", "Ngày đóng", "Ghi chú"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Data
    status_map = {
        'pending': 'Chờ đóng',
        'overdue': 'Quá hạn',
        'verified': 'Đã đóng',
        'cancelled': 'Đã hủy'
    }
    
    method_map = {
        'cash': 'Tiền mặt',
        'transfer': 'Chuyển khoản',
        'auto': 'Tự động'
    }
    
    total_amount = 0
    paid_amount = 0
    
    for idx, payment in enumerate(payments_data, 1):
        row = idx + 4
        amount = payment.get('amount', 0)
        total_amount += amount
        
        status = payment.get('payment_status', 'pending')
        if status == 'verified':
            paid_amount += amount
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=payment.get('member_name', ''))
        ws.cell(row=row, column=3, value=format_currency(amount))
        ws.cell(row=row, column=4, value=payment.get('reference_code', ''))
        ws.cell(row=row, column=5, value=method_map.get(payment.get('payment_method', 'transfer'), 'Chuyển khoản'))
        ws.cell(row=row, column=6, value=status_map.get(status, status))
        ws.cell(row=row, column=7, value=format_date(payment.get('due_date')))
        ws.cell(row=row, column=8, value=format_date(payment.get('paid_at')))
        ws.cell(row=row, column=9, value=payment.get('notes', ''))
        
        # Color coding for status
        status_cell = ws.cell(row=row, column=6)
        if status == 'verified':
            status_cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        elif status == 'overdue':
            status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        elif status == 'pending':
            status_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        for col in range(1, 10):
            apply_data_style(ws.cell(row=row, column=col), is_currency=(col == 3))
    
    # Summary row
    summary_row = len(payments_data) + 6
    ws.cell(row=summary_row, column=1, value="TỔNG CỘNG")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=format_currency(total_amount))
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    
    ws.cell(row=summary_row + 1, column=1, value="ĐÃ THU")
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True, color="16A34A")
    ws.cell(row=summary_row + 1, column=3, value=format_currency(paid_amount))
    ws.cell(row=summary_row + 1, column=3).font = Font(bold=True, color="16A34A")
    
    ws.cell(row=summary_row + 2, column=1, value="CÒN THIẾU")
    ws.cell(row=summary_row + 2, column=1).font = Font(bold=True, color="DC2626")
    ws.cell(row=summary_row + 2, column=3, value=format_currency(total_amount - paid_amount))
    ws.cell(row=summary_row + 2, column=3).font = Font(bold=True, color="DC2626")
    
    auto_adjust_columns(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_transactions_excel(transactions_data):
    """
    Generate Excel file for webhook transactions
    
    Args:
        transactions_data: List of transaction dictionaries
    
    Returns:
        BytesIO buffer containing Excel file
    """
    wb = create_styled_workbook()
    ws = wb.active
    ws.title = "Giao dịch"
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "BÁO CÁO GIAO DỊCH WEBHOOK"
    ws['A1'].font = Font(bold=True, size=14, color="0369A1")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Export date
    ws['A2'] = f"Xuất ngày: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="64748B")
    
    # Headers
    headers = ["STT", "Mã GD", "Số tiền", "Nội dung CK", "Tài khoản", "Thời gian", "Trạng thái", "Thanh toán liên kết", "Lỗi", "Ngày tạo"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)
    
    # Data
    status_map = {
        'success': 'Thành công',
        'failed': 'Thất bại',
        'pending_review': 'Cần xem xét',
        'rejected': 'Từ chối'
    }
    
    for idx, txn in enumerate(transactions_data, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=txn.get('external_id', ''))
        ws.cell(row=row, column=3, value=format_currency(txn.get('amount', 0)))
        ws.cell(row=row, column=4, value=txn.get('content', ''))
        ws.cell(row=row, column=5, value=txn.get('account_number', ''))
        ws.cell(row=row, column=6, value=format_date(txn.get('transaction_date')))
        ws.cell(row=row, column=7, value=status_map.get(txn.get('status', ''), txn.get('status', '')))
        ws.cell(row=row, column=8, value=txn.get('payment_id', '') or 'Chưa liên kết')
        ws.cell(row=row, column=9, value=txn.get('error_message', ''))
        ws.cell(row=row, column=10, value=format_date(txn.get('created_at')))
        
        for col in range(1, 11):
            apply_data_style(ws.cell(row=row, column=col))
    
    auto_adjust_columns(ws)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_hui_group_report_excel(hui_group_data, schedules_data, members_data):
    """
    Generate comprehensive Excel report for a hui group
    
    Args:
        hui_group_data: Hui group info dictionary
        schedules_data: List of schedule dictionaries
        members_data: List of member dictionaries
    
    Returns:
        BytesIO buffer containing Excel file
    """
    wb = create_styled_workbook()
    
    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Tổng quan"
    
    ws1['A1'] = f"BÁO CÁO DÂY HỤI: {hui_group_data.get('name', '')}"
    ws1['A1'].font = Font(bold=True, size=16, color="0369A1")
    ws1.merge_cells('A1:D1')
    
    ws1['A3'] = "Xuất ngày:"
    ws1['B3'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    ws1['A5'] = "THÔNG TIN DÂY HỤI"
    ws1['A5'].font = Font(bold=True, size=12)
    
    info_rows = [
        ("Tên dây hụi:", hui_group_data.get('name', '')),
        ("Chu kỳ:", hui_group_data.get('cycle_type', '')),
        ("Số tiền/kỳ:", format_currency(hui_group_data.get('amount_per_cycle', 0))),
        ("Tổng số kỳ:", hui_group_data.get('total_cycles', 0)),
        ("Kỳ hiện tại:", hui_group_data.get('current_cycle', 1)),
        ("Số thành viên:", hui_group_data.get('total_members', 0)),
        ("Ngày bắt đầu:", format_date(hui_group_data.get('start_date'))),
        ("Trạng thái:", 'Đang hoạt động' if hui_group_data.get('is_active') else 'Tạm dừng'),
    ]
    
    for idx, (label, value) in enumerate(info_rows, 7):
        ws1[f'A{idx}'] = label
        ws1[f'A{idx}'].font = Font(color="64748B")
        ws1[f'B{idx}'] = value
        ws1[f'B{idx}'].font = Font(bold=True)
    
    # Sheet 2: Members
    ws2 = wb.create_sheet("Thành viên")
    headers = ["STT", "Họ tên", "Số điện thoại", "Số chân", "Điểm tín dụng", "Trạng thái"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    for idx, member in enumerate(members_data, 1):
        ws2.cell(row=idx+1, column=1, value=idx)
        ws2.cell(row=idx+1, column=2, value=member.get('name', ''))
        ws2.cell(row=idx+1, column=3, value=member.get('phone', ''))
        ws2.cell(row=idx+1, column=4, value=member.get('slot_count', 1))
        ws2.cell(row=idx+1, column=5, value=member.get('credit_score', 100))
        ws2.cell(row=idx+1, column=6, value='Hoạt động' if member.get('is_active', True) else 'Ngừng')
    
    auto_adjust_columns(ws2)
    
    # Sheet 3: Schedules
    ws3 = wb.create_sheet("Lịch kỳ")
    headers = ["Kỳ", "Ngày hạn", "Người hốt", "Tổng thu", "Phí chủ hụi", "Trạng thái"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        apply_header_style(cell)
    
    for idx, schedule in enumerate(schedules_data, 1):
        ws3.cell(row=idx+1, column=1, value=schedule.get('cycle_number', idx))
        ws3.cell(row=idx+1, column=2, value=format_date(schedule.get('due_date')))
        ws3.cell(row=idx+1, column=3, value=schedule.get('receiver_name', 'Chưa chỉ định'))
        ws3.cell(row=idx+1, column=4, value=format_currency(schedule.get('total_amount', 0)))
        ws3.cell(row=idx+1, column=5, value=format_currency(schedule.get('owner_fee', 0)))
        ws3.cell(row=idx+1, column=6, value='Hoàn thành' if schedule.get('is_completed') else 'Chưa hoàn thành')
    
    auto_adjust_columns(ws3)
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
